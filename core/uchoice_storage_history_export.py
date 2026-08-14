"""
Excel export of the FULL storage-movement history behind a view_storage_history
query -- core/result_message.py's chat reply caps the detail list to the
latest 10 movements (WeCom Kefu's hard 2048-UTF-8-byte send_text limit made
an uncapped reply a 100% silent delivery failure), so this is the only way
to see everything for a range with real activity. Row-selection logic is
factored out here and reused by handlers/uchoice/queries.py's
QueryStorageHistoryHandler, so the chat summary and the export can never
silently drift apart -- same principle core/uchoice_invoice.py already
applies to the invoice/export pair.
"""
import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from sqlalchemy.orm import Session as DBSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from core.xlsx_determinism import freeze_xlsx_timestamps

_HEADER_FONT = Font(bold=True)
_TXN_TYPE_LABELS = {
    "inbound": "入库", "outbound": "出库",
    "convert_in": "转换入", "convert_out": "转换出",
    "move_in": "调拨入", "move_out": "调拨出",
    "transfer_in": "转仓入", "transfer_out": "转仓出",
    "adjust": "调整", "recount": "盘点",
}


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 60)


def query_storage_history_rows(
    db: DBSession, warehouse_code: str, start_month: str, end_month: str,
) -> tuple[list, date, date]:
    """
    Shared with handlers/uchoice/queries.py's QueryStorageHistoryHandler --
    identical range-resolution (capped at today for the current,
    still-in-progress month) and row shape, so the chat summary and the
    export always agree on what "this range" means.
    """
    import calendar
    from models.uchoice import UchoiceStorageTxn

    start_year, start_mo = (int(p) for p in start_month.split("-"))
    end_year, end_mo = (int(p) for p in end_month.split("-"))
    start = date(start_year, start_mo, 1)
    month_end = date(end_year, end_mo, calendar.monthrange(end_year, end_mo)[1])

    today = datetime.now(timezone.utc).date()
    range_end = min(month_end, today) if (end_year, end_mo) == (today.year, today.month) else month_end
    end_exclusive = range_end + timedelta(days=1)

    rows = (
        db.query(UchoiceStorageTxn)
        .filter(
            UchoiceStorageTxn.warehouse_code == warehouse_code,
            UchoiceStorageTxn.created_at >= start,
            UchoiceStorageTxn.created_at < end_exclusive,
        )
        .order_by(UchoiceStorageTxn.created_at)
        .all()
    )
    return rows, start, range_end


def build_storage_history_workbook(
    db: DBSession, warehouse_code: str, start_month: str, end_month: str | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """
    generated_at: see core/uchoice_invoice_export.py's build_invoice_workbook
    docstring -- same determinism requirement, same fix (freeze_xlsx_timestamps).
    """
    from core.uchoice_context import sku_label_map

    end_month = end_month or start_month
    generated_at = generated_at or datetime.now(timezone.utc)
    rows, start, range_end = query_storage_history_rows(db, warehouse_code, start_month, end_month)
    sku_labels = sku_label_map(db)

    wb = Workbook()
    wb.properties.created = generated_at
    wb.properties.modified = generated_at

    ws = wb.active
    ws.title = "Detail"
    ws.append(["Warehouse", warehouse_code])
    ws.append(["Range", f"{start.isoformat()} to {range_end.isoformat()}"])
    ws.append(["Generated at (UTC)", generated_at.strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    header_row = ws.max_row + 1
    _write_header(ws, header_row, ["Date", "Type", "SKU", "Boxes per Pallet", "Pallet Delta"])
    net_by_sku: dict[str, int] = {}
    for r in rows:
        sku_label = sku_labels.get(r.sku_code, r.sku_code)
        ws.append([
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            _TXN_TYPE_LABELS.get(r.txn_type, r.txn_type),
            sku_label,
            r.boxes_per_pallet,
            r.pallet_delta,
        ])
        net_by_sku[sku_label] = net_by_sku.get(sku_label, 0) + r.pallet_delta
    _autosize(ws)

    ws2 = wb.create_sheet("Net Change by SKU")
    _write_header(ws2, 1, ["SKU", "Net Pallet Change"])
    for sku_label, net in sorted(net_by_sku.items()):
        ws2.append([sku_label, net])
    total_row = ws2.max_row + 1
    ws2.append(["Total", sum(net_by_sku.values())])
    ws2.cell(row=total_row, column=1).font = _HEADER_FONT
    ws2.cell(row=total_row, column=2).font = _HEADER_FONT
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return freeze_xlsx_timestamps(buf.getvalue(), generated_at)


def build_storage_history_artifact(
    db: DBSession, warehouse_code: str, start_month: str, end_month: str | None, request_log_id,
) -> dict:
    """
    Channel-neutral artifact wrapper, matching handlers/uchoice/pdf_stub.py's
    and core/uchoice_invoice_export.py's {bytes, filename, content_type,
    artifact_key} shape -- so Kefu delivery (core/kefu_delivery.py's
    enqueue_file) and its replay path (core/kefu_artifact_loader.py) handle
    this exactly like any other durable Kefu file.

    generated_at is read from the persisted RequestLog.created_at, same
    reasoning as build_invoice_artifact: stable across retries/regeneration.
    """
    from models.request_log import RequestLog

    end_month = end_month or start_month
    log = db.query(RequestLog).filter_by(log_id=request_log_id).first() if request_log_id else None
    generated_at = log.created_at if log is not None else None
    data = build_storage_history_workbook(db, warehouse_code, start_month, end_month, generated_at=generated_at)
    return {
        "bytes": data,
        "filename": f"storage_history_{warehouse_code}_{start_month}_{end_month}.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "artifact_key": f"{request_log_id}:storage_history_workbook",
    }
