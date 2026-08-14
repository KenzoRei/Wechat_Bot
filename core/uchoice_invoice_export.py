"""
Excel export of the detail data behind an invoice — one row per contributing
transaction, not just the aggregated totals compute_invoice() returns. Uses
the exact same row-selection helpers as compute_invoice() (core/uchoice_invoice.py)
so the workbook and the chat summary can never silently drift apart.
"""
import io
import re
import zipfile
from datetime import datetime, timezone
from decimal import Decimal
from sqlalchemy.orm import Session as DBSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from core.uchoice_invoice import compute_invoice, _resolve_range, _completed_logs, _ledger_rows
from core.uchoice_context import sku_label_map, get_original_fields

_HEADER_FONT = Font(bold=True)
_CORE_XML_MODIFIED_RE = re.compile(
    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)"
)


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 60)


def _freeze_xlsx_timestamps(data: bytes, generated_at: datetime) -> bytes:
    """
    openpyxl's Workbook.save() unconditionally re-stamps
    properties.modified with datetime.now() during the save itself --
    setting wb.properties.modified beforehand (at any point, including
    immediately before save()) has no effect, confirmed empirically.
    docProps/core.xml's <dcterms:modified> is the only content difference
    between two saves of an otherwise-identical workbook; every ZIP
    entry's own date_time metadata already matches at the granularity
    that matters here. Rewrite that one XML value post-save and re-zip
    with every entry's date_time pinned to generated_at, so two builds of
    the same underlying data are byte-identical regardless of when each
    was actually generated -- required for Kefu's durable delivery queue,
    which verifies a content hash before every send (core/kefu_delivery.py).
    """
    stamp = generated_at.strftime("%Y-%m-%dT%H:%M:%SZ").encode("ascii")
    date_time = (generated_at.year, generated_at.month, generated_at.day,
                 generated_at.hour, generated_at.minute, generated_at.second)

    src = zipfile.ZipFile(io.BytesIO(data))
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as out:
        for info in src.infolist():
            content = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                content = _CORE_XML_MODIFIED_RE.sub(rb"\g<1>" + stamp + rb"\g<2>", content)
            new_info = zipfile.ZipInfo(info.filename, date_time=date_time)
            new_info.compress_type = zipfile.ZIP_DEFLATED
            out.writestr(new_info, content)
    return out_buf.getvalue()


def build_invoice_workbook(
    db: DBSession, warehouse_code: str, start_month: str, end_month: str | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """
    generated_at: pass a stable, persisted timestamp (e.g. RequestLog.created_at)
    for any caller that needs byte-identical regeneration -- Kefu's durable
    delivery queue verifies a content hash before every send (core/kefu_delivery.py),
    so a wall-clock default here would make every retry/redelivery mismatch
    and fail permanently, the same non-idempotent trap
    handlers/uchoice/pdf_stub.py's delivery_date already avoids via
    RequestLog.created_at. Also fixes the workbook's own created/modified
    properties to the same value -- openpyxl stamps those with datetime.now()
    by default on every save, which alone is enough to make two otherwise-
    identical workbooks hash differently.
    Defaults to datetime.now() for Smart Robot's one-shot, never-re-verified
    send path, where determinism doesn't matter.
    """
    end_month = end_month or start_month
    start, end, end_exclusive = _resolve_range(start_month, end_month)
    summary = compute_invoice(db, warehouse_code, start_month, end_month)
    sku_labels = sku_label_map(db)
    generated_at = generated_at or datetime.now(timezone.utc)

    wb = Workbook()
    wb.properties.created = generated_at
    wb.properties.modified = generated_at

    # ── Summary ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Warehouse", warehouse_code])
    ws.append(["Range", f"{start_month} to {end_month}"])
    ws.append(["Generated at (UTC)", generated_at.strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    _write_header(ws, ws.max_row + 1, ["Charge", "Amount (USD)"])
    ws.append(["Transportation fee", float(summary["transportation_fee"])])
    ws.append(["Palletization fee", float(summary["palletization_fee"])])
    ws.append(["Unpacking fee", float(summary["unpacking_fee"])])
    ws.append(["Storage fee", float(summary["storage_fee"])])
    total_row = ws.max_row + 1
    ws.append(["Total", float(summary["total"])])
    ws.cell(row=total_row, column=1).font = _HEADER_FONT
    ws.cell(row=total_row, column=2).font = _HEADER_FONT
    _autosize(ws)

    # ── Transportation & Palletization (outbound completions) ──────────────
    from models.uchoice import UchoiceAddress

    ws2 = wb.create_sheet("outbound")
    _write_header(ws2, 1, [
        "Serial Number", "Completed At (UTC)", "SKU Lines",
        "Destination Company", "Destination Address",
        "Transportation Fee", "Palletization Fee",
    ])
    outbound_logs = _completed_logs(db, "uchoice_outbound_request", warehouse_code, start, end_exclusive)
    for log in outbound_logs:
        result = log.result or {}
        lines = result.get("fulfillment_lines") or []
        sku_summary = "; ".join(
            f"{sku_labels.get(l.get('sku_code'), l.get('sku_code', '?'))} x{l.get('pallet_count', l.get('box_count', '?'))}"
            for l in lines
        )

        # destination isn't in result — it's on the original request, not the
        # completion's own fields, so it's resolved the same way the
        # confirmation/response builders do (core/uchoice_context.py).
        destination_company = ""
        destination_addr = ""
        original_fields = get_original_fields(db, log)
        destination_address_id = original_fields.get("destination_address_id")
        if destination_address_id:
            addr = db.query(UchoiceAddress).filter_by(address_id=destination_address_id).first()
            if addr:
                destination_company = addr.company_name or ""
                destination_addr = addr.addr

        ws2.append([
            log.serial_number,
            log.completed_at.strftime("%Y-%m-%d %H:%M") if log.completed_at else "",
            sku_summary,
            destination_company,
            destination_addr,
            float(Decimal(str(result.get("transportation_fee", 0)))),
            float(Decimal(str(result.get("palletization_fee", 0)))),
        ])
    _autosize(ws2)

    # ── Unpacking (inbound completions) ─────────────────────────────────────
    ws3 = wb.create_sheet("inbound")
    _write_header(ws3, 1, ["Serial Number", "Completed At (UTC)", "SKU Lines", "Unpacking Fee"])
    inbound_logs = _completed_logs(db, "uchoice_inbound_request", warehouse_code, start, end_exclusive)
    for log in inbound_logs:
        result = log.result or {}
        lines = result.get("received_lines") or []
        sku_summary = "; ".join(
            f"{sku_labels.get(l.get('sku_code'), l.get('sku_code', '?'))} x{l.get('pallet_count', l.get('box_count', '?'))}"
            for l in lines
        )
        ws3.append([
            log.serial_number,
            log.completed_at.strftime("%Y-%m-%d %H:%M") if log.completed_at else "",
            sku_summary,
            float(Decimal(str(result.get("unpacking_fee", 0)))),
        ])
    _autosize(ws3)

    # ── Storage (daily ledger) ───────────────────────────────────────────
    ws4 = wb.create_sheet("Storage")
    _write_header(ws4, 1, ["Date", "Pallet Count", "Storage Fee"])
    for row in _ledger_rows(db, warehouse_code, start, end):
        ws4.append([row.fee_date.isoformat(), row.pallet_count, float(row.storage_fee)])
    _autosize(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return _freeze_xlsx_timestamps(buf.getvalue(), generated_at)


def build_invoice_artifact(
    db: DBSession, warehouse_code: str, start_month: str, end_month: str | None, request_log_id
) -> dict:
    """
    Channel-neutral artifact wrapper around build_invoice_workbook, matching
    the {bytes, filename, content_type, artifact_key} shape handlers/uchoice/
    pdf_stub.py's PDF artifacts use -- so Kefu delivery (core/kefu_delivery.py's
    enqueue_file) and its replay path (core/kefu_artifact_loader.py) can
    handle an invoice workbook exactly like any other durable Kefu file, no
    Excel-specific casing needed there. artifact_key is stable per (request,
    doc_type) for the same idempotent-regeneration reason PDFs use it.

    generated_at is read from the persisted RequestLog.created_at, mirroring
    handlers/uchoice/pdf_stub.py's delivery_date -- both call sites
    (core/kefu_turn_apply.py's initial build, core/kefu_artifact_loader.py's
    later regeneration) pass a request_log_id whose row already exists by
    construction, so this stays stable across retries/regeneration.
    """
    from models.request_log import RequestLog

    end_month = end_month or start_month
    log = db.query(RequestLog).filter_by(log_id=request_log_id).first() if request_log_id else None
    generated_at = log.created_at if log is not None else None
    data = build_invoice_workbook(db, warehouse_code, start_month, end_month, generated_at=generated_at)
    return {
        "bytes": data,
        "filename": f"invoice_{warehouse_code}_{start_month}_{end_month}.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "artifact_key": f"{request_log_id}:invoice_workbook",
    }
